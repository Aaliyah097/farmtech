import random
import datetime
from string import ascii_letters

import django.db.utils
import pandas as pd
import openpyxl
from django.apps import AppConfig

from farmtech import settings


def random_password(strength: int = 8) -> str:
    return "".join(random.choice(ascii_letters) for _ in range(strength))


class UsersConfig(AppConfig):
    default_auto_field = "django.db.models.BigAutoField"
    name = "src.users"
    verbose_name = "Пользователи"

    def ready(self):
        return  # TODO
        from django.core.management import call_command

        call_command("migrate", "users")

        self.load_users_jobs()
        self.load_departments()
        self.load_jobs()
        self.load_users_departments()

        self.load_users_employment_date()

        self.load_phone_numbers()

    def load_phone_numbers(self):
        from src.users.models import User

        wb = openpyxl.load_workbook('Номера телефонов.xlsx')
        sheet = wb.active

        for row in range(5, sheet.max_row):
            fio = sheet.cell(row, 2).value

            if not fio:
                continue
            try:
                last_name, first_name, middle_name = fio.split(" ")
            except (ValueError, AttributeError):
                continue

            user = User.objects.filter(
                first_name__iexact=first_name.lower().strip(),
                last_name__iexact=last_name.lower().strip(),
                middle_name__iexact=middle_name.lower().strip()
            ).first()

            if not user:
                continue

            phone = sheet.cell(row, 1).value
            if not phone:
                continue

            user.phone = str(phone)
            user.save()

    def load_users_employment_date(self):
        from src.users.models import User

        wb = openpyxl.load_workbook('Сотрудники_дата_приема_город_15_04_2024.xlsx')
        sheet = wb.active

        for row in range(7, sheet.max_row):
            fio = sheet.cell(row, 2).value

            if not fio:
                continue
            try:
                last_name, first_name, middle_name = fio.split(" ")
            except (ValueError, AttributeError):
                continue

            user = User.objects.filter(
                first_name__iexact=first_name.lower().strip(),
                last_name__iexact=last_name.lower().strip(),
                middle_name__iexact=middle_name.lower().strip()
            ).first()

            if not user:
                continue

            employment_date = sheet.cell(row, 5).value
            employment_date = datetime.datetime.strptime(employment_date, '%d.%m.%Y')

            user.employment_date = employment_date
            user.save()

    def load_users_departments(self):
        from src.users.models import Departments, User, Jobs

        file = pd.read_csv("users.csv")
        for idx, row in file.iterrows():
            print('users.csv', idx)
            last_name, first_name, middle_name, job_name, fio, email, department_name = (
                row['Фамилия'],
                row['Имя'],
                row['Отчетство'],
                row['Должность'],
                row['ФИО'],
                row['Email'],
                row['Подразделение'],
            )

            job, _ = Jobs.objects.get_or_create(name=job_name)
            dep, _ = Departments.objects.get_or_create(name=department_name)

            try:
                user = User.objects.get(email=email)
            except User.DoesNotExist:
                continue

            user.departments.add(dep)

            user.save()

    def load_jobs(self):
        from src.users.models import Jobs

        file = pd.read_csv("jobs.csv")
        for idx, row in file.iterrows():
            print("jobs.csv", idx)
            name = row['Название']

            try:
                job = Jobs.objects.get(name=name)
            except Jobs.DoesNotExist:
                job = Jobs(name=name)
                job.save()

    def load_departments(self):
        from src.users.models import Departments, User

        file = pd.read_csv("departments.csv")
        for idx, row in file.iterrows():
            print("departments.csv", idx)
            name, _, manager_email = (
                row['Название'],
                row['Начальник ФИО'],
                row['Начальник почта']
            )
            try:
                dep = Departments.objects.get(name=name)
            except Departments.DoesNotExist:
                dep = None

            try:
                manager = User.objects.get(email=manager_email)
            except User.DoesNotExist:
                manager = None

            if dep:
                dep.manager = manager
            else:
                dep = Departments(
                    name=name,
                    manager=manager
                )
            dep.save()

    def load_users_jobs(self):
        from src.users.models import Jobs, User

        file = pd.read_csv("users_jobs.csv")
        for idx, row in file.iterrows():
            print('users_jobs.csv', idx)
            job, fio, organization, email = (
                row["Должность"],
                row["ФИО"],
                row["Организация"],
                row["Почта"],
            )
            split_fio = fio.split(" ")
            try:
                first_name = split_fio[1]
            except IndexError:
                first_name = None
            try:
                last_name = split_fio[0]
            except IndexError:
                last_name = None
            try:
                middle_name = split_fio[2]
            except IndexError:
                middle_name = None

            job, _ = Jobs.objects.get_or_create(name=job)
            username = email.split("@")[0]
            try:
                User.objects.get(username=username)
                User.objects.get(email=email)
            except User.DoesNotExist:
                pass
            else:
                continue
            try:
                user, is_created = User.objects.get_or_create(
                    username=username,
                    email=email,
                )
            except django.db.utils.IntegrityError:
                continue

            if is_created:
                user.set_password(random_password())
                user.first_name = first_name
                user.last_name = last_name
                user.middle_name = middle_name
                user.job = job
                user.company = organization
                user.save()
