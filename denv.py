import os
import sys
import openpyxl
import datetime
import random
from string import ascii_letters
import pandas as pd


def random_password(strength: int = 8) -> str:
    return "".join(random.choice(ascii_letters) for _ in range(strength))

def time_estimation(func):
    def _wrapper(*args, **kwargs):
        start = datetime.datetime.now()
        result = func(*args, **kwargs)
        print(f'{func.__name__} -> Цикл завершен, с:',
                (datetime.datetime.now() - start).total_seconds())
        return result

    return _wrapper

sys.path.append(
    os.path.join(os.path.dirname(__file__), 'farmtech')
)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "farmtech.settings")

import django
django.setup()
print(os.environ.get("POSTGRES_HOST"))
from django.core.management import call_command

from src.users.models import *
from src.reports.models import *
from src.orders.models import *
from src.news.models import *
from src.meetings.models import *
from src.auth2.models import *
from src.auth2.signup.service import SignUpService
from src.users.users.repo import UsersRepository


class UsersConfig:
    def ready(self):
        call_command("migrate", "users")
        
        self.load_regions()
        self.load_users()
        self.load_users()
        self.load_users_departments_jobs()
        self.load_user_departments_cities()
        self.load_users_employment_date__jobs__departments()
        self.load_users_phones()

    def email_to_lowercase(self):
        users = User.objects.all()
        print(f'users -> {users.count()}')
        for user in users:
            user_email = UsersRepository.validate_email(user.email)
            if user_email != user.email:
                ex_users = User.objects.filter(email__iexact=user_email)
                if ex_users.count() > 1:
                    user.delete()
                    print(f"duplicate -> {ex_users}")
                    continue
                user.email = user_email
                user.save()

    def load_regions(self):
        # Регионы_корректно.xlsx
        """
        Регион
        """
        wb = openpyxl.load_workbook('Регионы_корректно.xlsx')
        sheet = wb.active

        for row in range(2, sheet.max_row):
            region, is_created = Regions.objects.get_or_create(name=sheet.cell(row, 1).value)
            if is_created:
                print(f"Создан новый регион -> {region.name}")

    def load_users_phones(self):
        # Номера_телефонов.xlsx.xlsx
        """
        Абонентский номер
        ФИО -> (поиск по ФИО, если не находит, то скипаем) (пользователя нельзя создать без почты)
        """
        wb = openpyxl.load_workbook('Номера_телефонов.xlsx.xlsx')
        sheet = wb.active

        for row in range(5, sheet.max_row):
            fio = sheet.cell(row, 2).value

            if not fio:
                continue
            try:
                last_name, first_name, middle_name = fio.split(" ")
            except (ValueError, AttributeError):
                continue

            user = User.objects.filter(
                first_name__iexact=first_name.lower().strip(),
                last_name__iexact=last_name.lower().strip(),
                middle_name__iexact=middle_name.lower().strip()
            ).first()

            if not user:
                continue

            phone = sheet.cell(row, 1).value
            if not phone:
                continue

            user.phone = str(phone)
            user.save()

    def load_users_employment_date__jobs__departments(self):
        # Сотрудники_дата_приема_город_15_04_2024.xlsx
        # Фармтек + Интелбио на одном листе
        """
        Подразделение ->
        Сотрудник ->
        Табельный номер
        Должность ->
        Дата приема
        Место работы (город)
        """
        wb = openpyxl.load_workbook('Сотрудники_дата_приема_город_15_04_2024.xlsx')
        sheet = wb.active

        for row in range(7, sheet.max_row):
            fio = sheet.cell(row, 2).value

            if not fio:
                continue
            try:
                last_name, first_name, middle_name = fio.split(" ")
            except (ValueError, AttributeError):
                continue

            user = User.objects.filter(
                first_name__iexact=first_name.lower().strip(),
                last_name__iexact=last_name.lower().strip(),
                middle_name__iexact=middle_name.lower().strip()
            ).first()

            if not user:
                continue

            employment_date = sheet.cell(row, 5).value
            employment_date = datetime.datetime.strptime(employment_date, '%d.%m.%Y')

            city = sheet.cell(row, 6).value

            department = sheet.cell(row, 7).value
            
            ex_department, is_created = Departments.objects.get_or_create(name=department)
            if is_created:
                print(f'Создан новый Департамент -> {department}')

            ex_department.staff.add(user)

            job = sheet.cell(row, 4).value
            job, is_created = Jobs.objects.get_or_create(name=job)
            if is_created:
                print(f'Создана новая Должность -> {job}')
            
            user.job = job
            user.employment_date = employment_date
            user.city = city
            user.save()

    def load_users(self):
        # юзеры ИБ.xlsx
        """
        Фамилия
        Имя
        Отчество
        Должность ->
        email ->
        Подразделение ->
        Руководитель ->
        """
        # юзеры ФТ.xlsx
        """
        Фамилия
        Имя
        Отчество
        Должность ->
        email ->
        Подразделение ->
        Руководитель ->
        """
        files = (
            ('юзеры ФТ.xlsx', "Фармтек"),
            ('юзеры ИБ.xlsx', 'Интелбио')
        )
        for file_name, company in files:
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active

            for row in range(2, sheet.max_row):
                last_name = sheet.cell(row, 1).value
                first_name = sheet.cell(row, 2).value
                middle_name = sheet.cell(row, 3).value
                
                job, is_created = Jobs.objects.get_or_create(name=sheet.cell(row, 4).value)
                if is_created:
                    print(f"Создана новая Должность -> {job.name}")

                email = sheet.cell(row, 5).value

                if not email:
                    continue

                login = email.split("@")[0]

                department, is_created = Departments.objects.get_or_create(name=sheet.cell(row, 6).value)
                if is_created:
                    print(f"Создан новый Департамент -> {department.name}")

                manager_fio = sheet.cell(row, 7).value
                try:
                    manager_last_name, manager_first_name, manager_middle_name = str(manager_fio).split(" ")
                    manager = User.objects.filter(
                        first_name__iexact=manager_first_name.lower().strip(),
                        last_name__iexact=manager_last_name.lower().strip(),
                        middle_name__iexact=manager_middle_name.lower().strip()
                    ).first()
                except (ValueError, AttributeError):
                    manager = None

                try:
                    user = User.objects.get(email=email)
                except User.DoesNotExist:
                    user = User.objects.create_user(
                        email=email, password=SignUpService.otp()
                    )
                    print(f'Создан новый Пользователь -> {user.email}')

                if manager:
                    user.manager = manager
                department.staff.add(user)
                user.job = job
                user.first_name = first_name
                user.last_name = last_name
                user.middle_name = middle_name
                user.company = company
                user.save()

    def load_user_departments_cities(self):
        # Сотрудники место работы (с городами).xlsx
        """
        Сотрудник (ФИО) ->
        Подразделение ->
        Должность ->
        Место работы (город)
        """
        wb = openpyxl.load_workbook("Сотрудники место работы (с городами).xlsx")
        sheet = wb.active

        for row in range(2, sheet.max_row):
            fio = sheet.cell(row, 2).value
            try:
                last_name, first_name, middle_name = str(fio).split(" ")
            except (ValueError, AttributeError):
                continue

            department, is_created = Departments.objects.get_or_create(name=sheet.cell(row, 3).value)
            if is_created:
                print(f"Создан новый Департамент -> {department.name}")

            job, is_created = Jobs.objects.get_or_create(name=sheet.cell(row, 4).value)
            if is_created:
                print(f"Создана новая Должность -> {job.name}")

            user = User.objects.filter(
                first_name__iexact=first_name.lower().strip(),
                last_name__iexact=last_name.lower().strip(),
                middle_name__iexact=middle_name.lower().strip()
            ).first()

            if not user:
                continue
            
            city = sheet.cell(row, 5).value
            user.city = city
            user.job = job
            department.staff.add(user)
            user.save()

    def load_users_departments_jobs(self):
        # сотрудники, департаменты, должности интелбио.xlsx
        """ Интелбио
        Подразделение ->
        Сотрудник ->
        Должность ->
        """
        wb = openpyxl.load_workbook("сотрудники, департаменты, должности интелбио.xlsx")
        sheet = wb['подразделения']

        for row in range(4, sheet.max_row):
            department, is_created = Departments.objects.get_or_create(
                name=sheet.cell(row, 1).value
            )
            if is_created:
                print(f"Создан новый Департамент -> {department.name}")

        sheet = wb['сотрудники']
        
        for row in range(12, sheet.max_row):
            fio = sheet.cell(row, 2).value
            try:
                last_name, first_name, middle_name = str(fio).split(" ")
            except (ValueError, AttributeError):
                continue

            job, is_created = Jobs.objects.get_or_create(name=sheet.cell(row, 6).value)
            if is_created:
                print(f"Создана новая Должность -> {job.name}")

            department, is_created = Departments.objects.get_or_create(name=sheet.cell(row, 11).value)
            if is_created:
                print(f"Создан новый Департамент -> {department.name}")

            user = User.objects.filter(
                first_name__iexact=first_name.lower().strip(),
                last_name__iexact=last_name.lower().strip(),
                middle_name__iexact=middle_name.lower().strip()
            ).first()

            if not user:
                continue

            user.job = job
            department.staff.add(user)
            user.save()

   
class NewsConfig:
    def ready(self):
        call_command("migrate", "news")

        file = pd.read_csv("trademarks.csv")
        for _, row in file.iterrows():
            description, mktu_cls, order_number, registration_number, owner, status, valid_until = (
                row['Словесное описание'],
                row['Классы МКТУ'],
                row['Номер заявки'],
                row['Номер регистрации'],
                row['Владелец'],
                'active',
                datetime.datetime.strptime(row['Действует до'], "%m/%d/%Y")
            )

            try:
                _ = Trademarks.objects.get(description=description)
                continue
            except Trademarks.DoesNotExist:
                pass

            trademark = Trademarks(
                description=description,
                mktu_cls=mktu_cls,
                order_number=order_number,
                registration_number=registration_number,
                owner=owner,
                status=status,
                valid_until=valid_until
            )
            trademark.save()


if __name__ == '__main__':
    c = UsersConfig()
